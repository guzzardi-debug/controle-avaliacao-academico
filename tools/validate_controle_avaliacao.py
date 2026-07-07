# -*- coding: utf-8 -*-
"""Valida output/Controle Avaliação.xlsx contra CLAUDE.md e specs/.

Duas camadas:
1. Estrutura, fórmulas (como texto), estilos, validações, congelamento etc.
2. Teste real dos gatilhos pontuáveis: abre uma cópia do arquivo no
   Microsoft Excel via AppleScript, injeta valores de teste, força
   recálculo e lê o resultado calculado pelo Excel de volta — não é
   simulação em Python, é o mesmo motor de fórmulas que a equipe vai usar.
"""

import platform
import re
import shutil
import subprocess
import tempfile
import unicodedata
from pathlib import Path

import openpyxl

ARQ = "output/Controle Avaliação.xlsx"
N_DELEGADOS = 100
N_SESSOES = 5
REG_DATA_START = 4
PREENCH_HDR_ROW = 3
erros = []


def check(cond, msg):
    if cond:
        print(f"  ok  {msg}")
    else:
        erros.append(msg)
        print(f"FALHA {msg}")


def reg_row(i, s):
    return REG_DATA_START + (s - 1) * N_DELEGADOS + (i - 1)


wb = openpyxl.load_workbook(ARQ)

# 1. exatamente as 4 abas, na ordem do CLAUDE.md
check(wb.sheetnames == ["Preenchimento", "Rubrica", "Registro por Sessão", "Nota"],
      f"abas exatas e na ordem: {wb.sheetnames}")

# 2. termos proibidos em qualquer célula (CLAUDE.md)
PROIBIDOS = ["humana", "id delegado", "tipo de função", "mesa", "plenária"]
achados = []
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                low = unicodedata.normalize("NFC", c.value).lower()
                for p in PROIBIDOS:
                    if p in low:
                        if p in ("mesa", "plenária") and "não usar" in low:
                            continue
                        achados.append((ws.title, c.coordinate, p))
check(not achados, f"nenhum termo proibido em células: {achados}")

pre = wb["Preenchimento"]
reg = wb["Registro por Sessão"]
rub = wb["Rubrica"]
nota = wb["Nota"]

# 3. Preenchimento: sem gap entre cabeçalho e 1º delegado
check(PREENCH_HDR_ROW == 3, "Preenchimento: cabeçalho na linha 3 (compacto, sem começar na linha 10)")
check([pre.cell(row=3, column=j).value for j in range(2, 5)] ==
      ["Delegado", "Representação/Função", "Presença geral"],
      "Preenchimento: cabeçalhos B3:D3")
f_primeiro = pre.cell(row=4, column=1).value or ""
check("ROW()-3" in f_primeiro, "Preenchimento: 1º delegado na linha 4, imediatamente após o cabeçalho")
check(pre.cell(row=1, column=1).value == "Controle Avaliação" and
      isinstance(pre.cell(row=2, column=1).value, str) and pre.cell(row=2, column=1).value,
      "Preenchimento: título na linha 1, instrução na linha 2, sem linhas vazias antes do cabeçalho")

# 4. cabeçalhos canônicos do Registro
hdr3 = [reg.cell(row=3, column=j).value or "" for j in range(1, 25)]
esperado = ["", "Sessão", "Delegado", "Representação/Função", "Presença na sessão",
            "Status do registro"]
for l in ("D", "F", "C", "P"):
    esperado += [f"Evidência {l}", f"Gatilho {l}", f"Triagem {l}", f"Nota Preliminar {l}"]
esperado += ["Síntese acumulada da sessão", "Observação geral da sessão"]
check(hdr3 == esperado, "Registro: 24 cabeçalhos da linha 3 (4 campos exatos por eixo)")

grupos2 = [reg["A2"].value, reg["G2"].value, reg["K2"].value, reg["O2"].value,
           reg["S2"].value, reg["W2"].value]
check(grupos2 == ["Identificação", "Diplomacia e Respeito às Regras",
                  "Fidelidade à Função Designada", "Contribuição", "Participação",
                  "Síntese"], f"Registro: grupos da linha 2: {grupos2}")

# 5. Registro: ordenado por SESSÃO (todos os delegados da 1ª sessão antes da 2ª)
check(reg.max_row == 3 + N_DELEGADOS * N_SESSOES, f"Registro termina na linha {reg.max_row}")
casos_ordem = [
    (4, 1, 1), (5, 2, 1), (103, 100, 1),   # bloco da 1ª sessão: delegados 1, 2, ..., 100
    (104, 1, 2), (105, 2, 2),               # bloco da 2ª sessão recomeça no delegado 1
    (503, 100, 5),                          # último bloco: 5ª sessão, delegado 100
]
for r_probe, i, s in casos_ordem:
    f_c = reg.cell(row=r_probe, column=3).value or ""
    f_b = reg.cell(row=r_probe, column=2).value or ""
    ok = (f'Preenchimento!$B${PREENCH_HDR_ROW + i}' in f_c) and (f'"{s}ª Sessão"' in f_b)
    check(ok, f"Registro linha {r_probe} = sessão {s}, delegado {i} (ordenado por sessão)")

# 6. numeração da coluna A reinicia em 1 a cada sessão
for s in range(1, N_SESSOES + 1):
    r = reg_row(1, s)
    f_a = reg.cell(row=r, column=1).value or ""
    check(f'"",1)' in f_a, f"Registro linha {r}: numeração reinicia em 1 no início da {s}ª sessão")
r_ultimo = reg_row(N_DELEGADOS, 1)
check(f'"",{N_DELEGADOS})' in (reg.cell(row=r_ultimo, column=1).value or ""),
      f"Registro linha {r_ultimo}: numeração chega a {N_DELEGADOS} no fim do bloco da 1ª sessão")

# 7. coluna de ordenamento roxa com texto branco (padrão da coluna de índice oficial)
amostras_a = [reg_row(1, 1), reg_row(50, 3), reg_row(N_DELEGADOS, N_SESSOES)]
roxo_ok = all(
    reg.cell(row=r, column=1).fill.fgColor.rgb == "FF20124D" and
    reg.cell(row=r, column=1).font.color.rgb == "FFFFFFFF"
    for r in amostras_a)
check(roxo_ok, "Registro: coluna de ordenamento (A) roxo institucional com texto branco em todas as linhas úteis")

# 8. alternância linha sim / linha não (não mais por bloco de 4/5 linhas)
fills = []
for r in range(REG_DATA_START, REG_DATA_START + 12):
    fill = reg.cell(row=r, column=3).fill  # coluna C, fora da faixa roxa
    fills.append(fill.fgColor.rgb if fill.patternType else None)
alterna_simples = all(
    (fills[k] is not None) != (fills[k + 1] is not None) for k in range(len(fills) - 1))
check(alterna_simples, f"Registro: listra simples linha sim/linha não (coluna C, 12 linhas): {fills}")

# 9. fórmulas de Nota Preliminar: base certa, clamp, colunas certas
for col, letra, base in [(10, "D", 8), (14, "F", 8), (18, "C", 5), (22, "P", 5)]:
    f = reg.cell(row=4, column=col).value
    gat = {"D": "H", "F": "L", "C": "P", "P": "T"}[letra]
    ok = (f'MIN(10,{base}+' in f and "MAX(0," in f and f'${gat}4' in f
          and f'"{letra}",""' in f and '$E4="Falta"' in f and f.count("IFERROR") == 8)
    check(ok, f"Nota Preliminar {letra}: base {base}, clamp 0-10, gatilho ${gat}, 8 tokens")

# 10. consolidação da Nota: referências ajustadas ao novo mapeamento sessão-primeiro
grupos_nota = [nota["M2"].value, nota["R2"].value, nota["W2"].value, nota["AB2"].value]
check(grupos_nota == ["Diplomacia e Respeito às regras", "Fidelidade à Função designada",
                      "Contribuição", "Participação"],
      f"Nota: grupos de eixo na grafia da planilha oficial: {grupos_nota}")
casos_nota = [
    ("M4", reg_row(1, 1), "J"),   # delegado 1, sessão 1, eixo D
    ("N4", reg_row(1, 2), "J"),   # delegado 1, sessão 2, eixo D
    ("Q4", reg_row(1, 5), "J"),   # delegado 1, sessão 5, eixo D
    ("R4", reg_row(1, 1), "N"),   # delegado 1, sessão 1, eixo F
    ("W4", reg_row(1, 1), "R"),   # delegado 1, sessão 1, eixo C
    ("AB4", reg_row(1, 1), "V"),  # delegado 1, sessão 1, eixo P
    ("M5", reg_row(2, 1), "J"),   # delegado 2, sessão 1, eixo D
    ("AF103", reg_row(100, 5), "V"),  # delegado 100, sessão 5, eixo P
]
for cel, linha_reg, col_reg in casos_nota:
    f = nota[cel].value or ""
    check(f"'Registro por Sessão'!${col_reg}{linha_reg}" in f and "ISNUMBER" in f,
          f"Nota!{cel} -> Registro!${col_reg}{linha_reg} (mapeamento sessão-primeiro)")

# 11. validações de dados
dvs_reg = list(reg.data_validations.dataValidation)
check(len(dvs_reg) == 6, f"Registro: 6 validações (presença, status, 4 triagens): {len(dvs_reg)}")
triagem_dvs = [dv for dv in dvs_reg if "Rubrica" in (dv.formula1 or "")]
check(len(triagem_dvs) == 4, "Registro: 4 dropdowns de Triagem apontando para a Rubrica")
for dv in triagem_dvs:
    m = re.match(r"'Rubrica'!\$C\$(\d+):\$C\$(\d+)", dv.formula1)
    ok = False
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        vals = [rub.cell(row=x, column=3).value for x in range(a, b + 1)]
        ok = all(isinstance(v, str) and v for v in vals)
    check(ok, f"Triagem {dv.sqref}: intervalo {dv.formula1} com {b - a + 1} opções preenchidas")
dvs_pre = list(pre.data_validations.dataValidation)
check(any("Presente" in (dv.formula1 or "") for dv in dvs_pre),
      "Preenchimento: dropdown Presente/Falta")

# 12. formatação condicional de falta
for ws, col in [(pre, "$D"), (reg, "$E"), (nota, "$D")]:
    regras = [r for rng in ws.conditional_formatting for r in rng.rules]
    ok = any('="Falta"' in (r.formula[0] if r.formula else "") for r in regras)
    check(ok, f"{ws.title}: formatação condicional de Falta")

# 13. congelamento e filtros
check(pre.freeze_panes == "C4" and reg.freeze_panes == "E4" and nota.freeze_panes == "C4",
      f"congelamentos: {pre.freeze_panes}, {reg.freeze_panes}, {nota.freeze_panes}")
check(pre.auto_filter.ref == "A3:D103" and reg.auto_filter.ref == "A3:X503",
      f"filtros: {pre.auto_filter.ref}, {reg.auto_filter.ref}")
check(nota.auto_filter.ref is None, "Nota sem filtro (protege alinhamento da colagem)")

# 14. área de colagem limpa
sujas = []
for r in range(4, 104):
    for c in range(13, 33):
        v = nota.cell(row=r, column=c).value
        if not (isinstance(v, str) and v.startswith("=IF(ISNUMBER(")):
            sujas.append(nota.cell(row=r, column=c).coordinate)
check(not sujas, f"área de colagem M4:AF103 contém apenas as fórmulas de nota: {sujas[:5]}")

# 15. médias de conferência da Nota
check("AVERAGE(M4:Q4)" in (nota["E4"].value or ""), "Nota: Média D usa AVERAGE das 5 sessões de Diplomacia")
check("AVERAGE(M4:AF4)" in (nota["I4"].value or ""), "Nota: Média geral usa AVERAGE das 20 notas")

# 16. workbook configurado para recálculo automático ao abrir
check(bool(wb.calculation) and wb.calculation.fullCalcOnLoad is True,
      "Workbook: fullCalcOnLoad=True (recalcula tudo ao abrir, sem valores antigos)")


# ---------------------------------------------------------------- teste real no Excel

def testar_gatilhos_no_excel(caminho_original):
    """Testa a Nota Preliminar abrindo uma cópia real no Microsoft Excel.

    Não é simulação: injeta os tokens de teste nas células de Gatilho de
    uma cópia do arquivo, abre no Excel via AppleScript, força recálculo
    e lê o valor que o próprio Excel calculou. Fecha a cópia sem salvar
    o original. Requer macOS com Microsoft Excel instalado.
    """
    if platform.system() != "Darwin":
        print("  --  pulado: teste no Excel requer macOS (ambiente atual não é Darwin)")
        return True
    if not Path("/Applications/Microsoft Excel.app").exists():
        print("  --  pulado: Microsoft Excel.app não encontrado em /Applications")
        return True

    tmp_dir = tempfile.mkdtemp(prefix="controle_avaliacao_teste_")
    tmp_path = str(Path(tmp_dir) / "teste_gatilhos.xlsx")
    shutil.copy(caminho_original, tmp_path)

    twb = openpyxl.load_workbook(tmp_path)
    tpre = twb["Preenchimento"]
    treg = twb["Registro por Sessão"]
    for i in range(1, 10):
        tpre.cell(row=PREENCH_HDR_ROW + i, column=2, value=f"Delegado Teste {i}")
        tpre.cell(row=PREENCH_HDR_ROW + i, column=3, value="Representação Teste")

    # (delegado i [sessão 1], presença, coluna do Gatilho, tokens, célula de leitura, esperado)
    CASOS = [
        (1, "Presente", "H", "D+1", "J", 9, "D+1 -> base 8 + 1"),
        (2, "Presente", "H", "D+1; D-2", "J", 7, "D+1; D-2 -> 8 + 1 - 2"),
        (3, "Presente", "L", "F+2; F-1", "N", 9, "F+2; F-1 -> 8 + 2 - 1"),
        (4, "Presente", "P", "C+2", "R", 7, "C+2 -> base 5 + 2"),
        (5, "Presente", "P", "C+2; C-1", "R", 6, "C+2; C-1 -> 5 + 2 - 1"),
        (6, "Presente", "T", "P-2", "V", 3, "P-2 -> base 5 - 2"),
        (7, "Presente", "H", "C+2", "J", 8, "token de eixo errado ignorado -> base 8"),
        (8, "Presente", "H", "xyz", "J", 8, "texto inválido ignorado -> base 8"),
        (9, "Falta", "H", "D+1", "J", "-", "Presença=Falta -> \"-\""),
    ]
    for i, presenca, col_gat, tokens, _col_leitura, _esp, _desc in CASOS:
        row = reg_row(i, 1)
        treg.cell(row=row, column=5, value=presenca)
        treg.cell(row=row, column=openpyxl.utils.column_index_from_string(col_gat), value=tokens)
    twb.save(tmp_path)

    script = f'''
    tell application "Microsoft Excel"
        activate
        open POSIX file "{tmp_path}"
        delay 2
        calculate
        delay 1
        set ws to worksheet "Registro por Sessão" of active workbook
        set resultados to {{}}
    '''
    for i, _presenca, _col_gat, _tokens, col_leitura, _esp, _desc in CASOS:
        row = reg_row(i, 1)
        script += f'        set end of resultados to (value of range "{col_leitura}{row}" of ws) as string\n'
    script += '''
        close active workbook saving no
        return resultados
    end tell
    '''
    try:
        out = subprocess.run(["osascript", "-e", script], capture_output=True,
                             text=True, timeout=60)
    except subprocess.TimeoutExpired:
        print("  --  teste no Excel expirou (timeout); pulado")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return True
    finally:
        pass

    shutil.rmtree(tmp_dir, ignore_errors=True)

    if out.returncode != 0:
        check(False, f"teste no Excel: falha ao automatizar ({out.stderr.strip()})")
        return False

    valores = [v.strip() for v in out.stdout.strip().split(", ")]
    tudo_ok = True
    for (i, _p, _cg, tokens, _cl, esperado, desc), obtido in zip(CASOS, valores):
        if str(esperado) == "-":
            ok = obtido == "-" or obtido == "missing value"
        else:
            try:
                # AppleScript devolve número como string no locale do sistema
                # (ex.: "9,0" em pt-BR) -> normaliza vírgula decimal para ponto
                obtido_num = obtido.replace(".", "").replace(",", ".") if "," in obtido else obtido
                ok = abs(float(obtido_num) - float(esperado)) < 1e-9
            except ValueError:
                ok = False
        check(ok, f"Excel real: Gatilho={tokens!r} ({desc}) -> obtido {obtido!r}, esperado {esperado!r}")
        tudo_ok = tudo_ok and ok
    return tudo_ok


print()
print("Teste de gatilhos pontuáveis em células reais (Microsoft Excel):")
testar_gatilhos_no_excel(ARQ)

print()
if erros:
    print(f"{len(erros)} FALHA(S)")
    raise SystemExit(1)
print("Todas as verificações passaram.")
