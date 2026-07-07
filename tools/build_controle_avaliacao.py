# -*- coding: utf-8 -*-
"""Gera output/Controle Avaliação.xlsx a partir das specs do repositório.

Fontes: CLAUDE.md, specs/MATRIZ_DE_AVALIAÇÃO.md, specs/ONTOLOGIA_ACADEMICA.md,
specs/ERRATA_GATILHOS_PONTUAVEIS.md, specs/PLANITA_NOTAS_CONTRATO_TECNICO.md.

Padrão visual reproduzido da Planilha de Notas oficial (input/):
título preto/branco 20pt, cabeçalhos roxo FF20124D com texto branco,
listras CCCCCC/branco, falta E6B8AF via formatação condicional,
bordas finas (preto entre blocos, B7B7B7 internas), Calibri.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- constantes

N_DELEGADOS = 100          # capacidade, igual à Planilha de Notas oficial
N_SESSOES = 5
REG_NAME = "Registro por Sessão"

PREENCH_HDR_ROW = 9        # cabeçalho do cadastro (dados a partir da linha 10)
REG_DATA_START = 4         # dados do Registro começam na linha 4
NOTA_DATA_START = 4        # dados da Nota começam na linha 4

ROXO = "FF20124D"
PRETO = "FF000000"
BRANCO = "FFFFFFFF"
CINZA_LISTRA = "FFCCCCCC"
CINZA_BORDA = "FFB7B7B7"
ROSA_FALTA = "FFE6B8AF"

F_TITULO = Font(name="Calibri", size=20, bold=True, color=BRANCO)
F_SUBTITULO = Font(name="Calibri", size=14, bold=True, color=BRANCO)
F_GRUPO = Font(name="Calibri", size=16, bold=True, color=BRANCO)
F_HDR = Font(name="Calibri", size=11, bold=True, color=BRANCO)
F_HDR12 = Font(name="Calibri", size=12, bold=True, color=BRANCO)
F_TXT = Font(name="Calibri", size=11)
F_TXT_B = Font(name="Calibri", size=11, bold=True)
F_TXT_I = Font(name="Calibri", size=11, italic=True, color="FF666666")

FILL_TITULO = PatternFill("solid", start_color=PRETO, end_color=PRETO)
FILL_HDR = PatternFill("solid", start_color=ROXO, end_color=ROXO)
FILL_LISTRA = PatternFill("solid", start_color=CINZA_LISTRA, end_color=CINZA_LISTRA)
FILL_FALTA = PatternFill("solid", start_color=ROSA_FALTA, end_color=ROSA_FALTA)

LADO_FINO = Side(style="thin", color=CINZA_BORDA)
LADO_PRETO = Side(style="thin", color=PRETO)
BORDA_FINA = Border(left=LADO_FINO, right=LADO_FINO, top=LADO_FINO, bottom=LADO_FINO)
BORDA_BLOCO = Border(left=LADO_PRETO, right=LADO_FINO, top=LADO_FINO, bottom=LADO_FINO)

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
AL_L_NW = Alignment(horizontal="left", vertical="center")

EIXOS = [
    # (letra, nome CLAUDE.md, nome oficial p/ área de colagem, base)
    ("D", "Diplomacia e Respeito às Regras", "Diplomacia e Respeito às regras", 8),
    ("F", "Fidelidade à Função Designada", "Fidelidade à Função designada", 8),
    ("C", "Contribuição", "Contribuição", 5),
    ("P", "Participação", "Participação", 5),
]

TRIAGEM = {
    "D": ["Procedimento adequado", "Uso correto de moção/questão de ordem",
          "Postura respeitosa em conflito", "Erro procedimental isolado",
          "Erro procedimental repetido", "Interrupção/fala fora de ordem",
          "Conflito improdutivo", "Conduta excludente/dominadora"],
    "F": ["Coerente com a função", "Coerência entre fala e documento",
          "Argumento compatível com interesses", "Concessão justificada",
          "Atuação genérica", "Proposta incompatível com função",
          "Mudança oportunista sem justificativa", "Incoerência em bloco/aliança"],
    "C": ["Sem contribuição observável", "Participa sem encaminhamento",
          "Proposta genérica", "Proposta aproveitável", "Proposta desenvolvida",
          "Cláusula/documento melhorado", "Organização de solução",
          "Destravamento de impasse", "Materialização decisiva",
          "Centralização excludente"],
    "P": ["Sem engajamento observável", "Presença irregular", "Fala pontual",
          "Discursos regulares", "Participa de negociação",
          "Interage com diferentes representações", "Presente em momento decisivo",
          "Acompanha documentos/debates", "Fala muito sem avanço"],
}

STATUS_REGISTRO = ["Pendente", "Em andamento", "Concluído", "Revisado"]

ESCALA = [
    (10, "Desempenho excepcional: evidência clara, impacto decisivo, aderência plena ao eixo"),
    (9, "Desempenho destacado: evidência concreta, impacto relevante"),
    (8, "Desempenho adequado ou bom, compatível com o esperado"),
    (7, "Desempenho aceitável, com pequenos limites ou irregularidades"),
    (6, "Desempenho limitado, irregular ou pouco consistente"),
    (5, "Desempenho mediano, mínimo ou insuficientemente demonstrado"),
    (4, "Problema relevante ou baixa atuação no eixo"),
    (3, "Problema grave, repetido ou impacto negativo claro"),
    (2, "Atuação muito insuficiente ou fortemente incompatível com o eixo"),
    (1, "Atuação quase inexistente ou conduta extremamente prejudicial"),
    (0, "Ausência total de atuação avaliável ou conduta incompatível com avaliação no eixo"),
]

MAX_TOKENS = 8
PAD = 30


# ---------------------------------------------------------------- fórmulas

def formula_nota_preliminar(letra, base, col_gatilho, row):
    """Nota Preliminar = limitar(base + soma dos tokens do eixo, 0, 10).

    Parser de tokens sem funções de matriz (compatível com Excel clássico,
    Excel 365 e Google Sheets): remove espaços e a letra do eixo, troca ';'
    por 30 espaços e lê até 8 fatias fixas com MID/TRIM/VALUE.
    Token inválido ou de outro eixo -> IFERROR -> 0.
    """
    limpo = (f'SUBSTITUTE(SUBSTITUTE(UPPER(${col_gatilho}{row})," ",""),"{letra}","")')
    pad = f'SUBSTITUTE({limpo},";",REPT(" ",{PAD}))'
    termos = []
    for k in range(MAX_TOKENS):
        off = k * PAD + 1
        termos.append(f'IFERROR(VALUE(TRIM(MID({pad},{off},{PAD}))),0)')
    soma = "+".join(termos)
    return (f'=IF($C{row}="","",IF($E{row}="Falta","-",'
            f'MAX(0,MIN(10,{base}+{soma}))))')


def simulate_tokens(text, letra):
    """Réplica em Python do parser da fórmula, para validação."""
    limpo = text.upper().replace(" ", "").replace(letra, "")
    pad = limpo.replace(";", " " * PAD)
    total = 0
    for k in range(MAX_TOKENS):
        fatia = pad[k * PAD: k * PAD + PAD].strip()
        try:
            total += int(fatia)
        except ValueError:
            pass
    return total


# ---------------------------------------------------------------- helpers

def stripe_fill(par):
    return FILL_LISTRA if par else None


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def title_block(ws, last_col, titulo, subtitulo):
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = titulo
    ws["A1"].font = F_TITULO
    ws["A1"].alignment = AL_L_NW
    ws.row_dimensions[1].height = 30
    for col in range(1, openpyxl.utils.column_index_from_string(last_col) + 1):
        ws.cell(row=1, column=col).fill = FILL_TITULO
    if subtitulo:
        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"] = subtitulo
        ws["A2"].font = F_SUBTITULO
        ws["A2"].alignment = AL_L_NW
        ws.row_dimensions[2].height = 22
        for col in range(1, openpyxl.utils.column_index_from_string(last_col) + 1):
            ws.cell(row=2, column=col).fill = FILL_TITULO


# ---------------------------------------------------------------- Preenchimento

def build_preenchimento(wb):
    ws = wb.active
    ws.title = "Preenchimento"
    ws.sheet_properties.tabColor = ROXO[2:]

    title_block(ws, "D", "Controle Avaliação", "Preenchimento — lista-mãe do comitê")

    instrucoes = [
        "Preencha um delegado por linha, a partir da linha 10. A ordem desta aba comanda o Registro por Sessão e a aba Nota.",
        "Delegado e Representação/Função são preenchidos pela equipe. Presença geral é a presença consolidada do evento.",
        "Não insira nem exclua linhas: as demais abas referenciam estas posições fixas.",
    ]
    for i, txt in enumerate(instrucoes):
        cell = ws.cell(row=4 + i, column=2, value=txt)
        cell.font = F_TXT_I
        cell.alignment = AL_L_NW

    headers = ["", "Delegado", "Representação/Função", "Presença geral"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=PREENCH_HDR_ROW, column=j, value=h)
        c.font = F_HDR12
        c.fill = FILL_HDR
        c.alignment = AL_C
        c.border = BORDA_FINA
    ws.row_dimensions[PREENCH_HDR_ROW].height = 20

    for i in range(1, N_DELEGADOS + 1):
        r = PREENCH_HDR_ROW + i
        ws.cell(row=r, column=1, value=f'=IF($B{r}="","",ROW()-{PREENCH_HDR_ROW})')
        fill = stripe_fill(i % 2 == 1)  # linha 10 (1º delegado) cinza, como a oficial
        for j in range(1, 5):
            c = ws.cell(row=r, column=j)
            c.font = F_TXT
            c.border = BORDA_FINA
            c.alignment = AL_L_NW if j in (2, 3) else Alignment(horizontal="center", vertical="center")
            if fill:
                c.fill = fill

    dv = DataValidation(type="list", formula1='"Presente,Falta"', allow_blank=True,
                        showErrorMessage=True,
                        error="Use apenas Presente ou Falta.")
    ws.add_data_validation(dv)
    dv.add(f"D{PREENCH_HDR_ROW + 1}:D{PREENCH_HDR_ROW + N_DELEGADOS}")

    ws.conditional_formatting.add(
        f"A{PREENCH_HDR_ROW + 1}:D{PREENCH_HDR_ROW + N_DELEGADOS}",
        FormulaRule(formula=[f'$D{PREENCH_HDR_ROW + 1}="Falta"'], fill=FILL_FALTA))

    ws.freeze_panes = f"C{PREENCH_HDR_ROW + 1}"
    ws.auto_filter.ref = f"A{PREENCH_HDR_ROW}:D{PREENCH_HDR_ROW + N_DELEGADOS}"
    set_widths(ws, {"A": 4, "B": 32, "C": 30, "D": 14})
    return ws


# ---------------------------------------------------------------- Rubrica

def build_rubrica(wb):
    ws = wb.create_sheet("Rubrica")
    ws.sheet_properties.tabColor = ROXO[2:]
    title_block(ws, "E", "Controle Avaliação", "Rubrica — referência operacional")
    set_widths(ws, {"A": 3, "B": 38, "C": 52, "D": 16, "E": 52})

    r = 4
    triagem_ranges = {}

    def secao(txt):
        nonlocal r
        ws.merge_cells(f"B{r}:E{r}")
        c = ws.cell(row=r, column=2, value=txt)
        c.font = F_GRUPO
        c.alignment = AL_L_NW
        for col in range(2, 6):
            ws.cell(row=r, column=col).fill = FILL_HDR
        ws.row_dimensions[r].height = 22
        r += 1

    def linha(txt, bold=False):
        nonlocal r
        ws.merge_cells(f"B{r}:E{r}")
        c = ws.cell(row=r, column=2, value=txt)
        c.font = F_TXT_B if bold else F_TXT
        c.alignment = AL_L
        r += 1

    def tabela(headers, rows, widths_cols=None):
        nonlocal r
        for j, h in enumerate(headers, start=2):
            c = ws.cell(row=r, column=j, value=h)
            c.font = F_HDR
            c.fill = FILL_HDR
            c.alignment = AL_C
            c.border = BORDA_FINA
        r += 1
        first = r
        for i, row_vals in enumerate(rows):
            fill = stripe_fill(i % 2 == 0)
            for j, v in enumerate(row_vals, start=2):
                c = ws.cell(row=r, column=j, value=v)
                c.font = F_TXT
                c.alignment = AL_L
                c.border = BORDA_FINA
                if fill:
                    c.fill = fill
            r += 1
        return first, r - 1

    # 1. Eixos e notas base
    secao("1. Eixos e notas base")
    tabela(["Eixo", "O que mede", "Nota base", "Pergunta avaliativa central"], [
        ("Diplomacia e Respeito às Regras",
         "Postura, decoro, domínio procedimental, relação com a presidência",
         8, "Atuou com postura adequada e respeito ao procedimento?"),
        ("Fidelidade à Função Designada",
         "Coerência com o papel atribuído (política externa para delegações; condução neutra para presidência)",
         8, "Foi coerente com o papel que recebeu?"),
        ("Contribuição",
         "Impacto e avanço material: proposta, cláusula, documento, acordo, destravamento, organização de solução, materialização",
         5, "Fez o comitê avançar concretamente?"),
        ("Participação",
         "Presença ativa, regularidade e volume de engajamento nas sessões",
         5, "Esteve presente e ativo nas sessões?"),
    ])
    linha("Base 8 (Diplomacia, Fidelidade): presunção de adequação — a nota se move para cima ou para baixo apenas mediante gatilho registrado.")
    linha("Base 5 (Contribuição, Participação): nada é presumido — a nota só sobe mediante atuação demonstrada e registrada.")
    linha("Participação alta não compensa Contribuição baixa. Os dois eixos são registrados de forma independente.", bold=True)
    r += 1

    # 2. Gatilhos pontuáveis
    secao("2. Gatilhos pontuáveis (campo Gatilho)")
    linha("Neste Controle Avaliação, o campo Gatilho usa tokens pontuáveis: cada token soma ou subtrai pontos diretamente na nota do eixo.", bold=True)
    linha("Os tokens não são os IDs fixos do catálogo da Matriz de Avaliação — ver specs/ERRATA_GATILHOS_PONTUAVEIS.md (decisão registrada do projeto).")
    linha("Formato: D+X / D-Y, F+X / F-Y, C+X / C-Y, P+X / P-Y, com X e Y inteiros. Validação recomendada: valores de -8 até +5.")
    linha("Vários tokens no mesmo campo, separados por ponto e vírgula (até 8 tokens por campo). Exemplo: C+2; C-1")
    linha("Nota Preliminar = nota base do eixo + soma dos tokens, sempre limitada entre 0 e 10.")
    linha("Tokens de outro eixo ou texto inválido são ignorados pela fórmula. Presença na sessão = Falta gera nota \"-\".")
    linha("Cada token exige Evidência correspondente no mesmo eixo e na mesma sessão. A fórmula organiza a pontuação; ela não substitui a análise nem a revisão acadêmica.", bold=True)
    tabela(["Exemplo de Gatilho", "Eixo", "Cálculo", "Nota Preliminar"], [
        ("C+2", "Contribuição", "5 + 2", 7),
        ("C+2; C-1", "Contribuição", "5 + 2 - 1", 6),
        ("D+1; D-2", "Diplomacia e Respeito às Regras", "8 + 1 - 2", 7),
        ("F+2; F-1", "Fidelidade à Função Designada", "8 + 2 - 1", 9),
        ("P-2", "Participação", "5 - 2", 3),
    ])
    r += 1

    # 3. Escala de interpretação
    secao("3. Escala de interpretação (0 a 10)")
    tabela(["Nota", "Interpretação", "", ""], [(n, t, "", "") for n, t in ESCALA])
    linha("A mesma nota tem significado diferente conforme a base do eixo: 5 em Contribuição é atuação mediana; 5 em Diplomacia é queda de 3 pontos abaixo da base.")
    r += 1

    # 4. Travas de registro
    secao("4. Travas de registro")
    linha("Nota 9 ou 10 exige evidência escrita com fato concreto, impacto claro e relação direta com o eixo.")
    linha("Nota abaixo da base do eixo exige evidência escrita.")
    linha("Participação não compensa Contribuição: volume de fala sem avanço material não eleva Contribuição.", bold=True)
    linha("Registro apenas observacional: preencher Evidência e Triagem, sem token no Gatilho (delta zero).")
    linha("A consolidação final e a premiação passam por revisão acadêmica. A Nota Preliminar não é nota final.")
    linha("Status do registro (Registro por Sessão): Pendente, Em andamento, Concluído, Revisado — controle operacional da equipe.")
    r += 1

    # 5. Triagem
    secao("5. Triagem — opções por eixo (menu suspenso)")
    linha("A Triagem classifica rapidamente o tipo de evidência para leitura acadêmica posterior. Ela não calcula nota: a pontuação vem apenas do campo Gatilho.", bold=True)
    for j, h in enumerate(["Eixo", "Opção de Triagem", "", ""], start=2):
        c = ws.cell(row=r, column=j, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = AL_C
        c.border = BORDA_FINA
    r += 1
    for letra, nome, _, _ in EIXOS:
        first = r
        for i, opcao in enumerate(TRIAGEM[letra]):
            ws.cell(row=r, column=3, value=opcao).font = F_TXT
            ws.cell(row=r, column=3).border = BORDA_FINA
            ws.cell(row=r, column=3).alignment = AL_L
            if i % 2 == 0:
                ws.cell(row=r, column=3).fill = FILL_LISTRA
            for col in (2, 4, 5):
                ws.cell(row=r, column=col).border = BORDA_FINA
            r += 1
        ws.merge_cells(start_row=first, start_column=2, end_row=r - 1, end_column=2)
        c = ws.cell(row=first, column=2, value=nome)
        c.font = F_TXT_B
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        triagem_ranges[letra] = f"'Rubrica'!$C${first}:$C${r - 1}"
    r += 1

    # 6. Padrão de evidência
    secao("6. Padrão de evidência")
    linha("Evidência observável: o quê, quem, quando, eixo e impacto. Pode ser mais longa quando necessário para a revisão acadêmica.")
    linha('Válido: "Redigiu cláusula aceita na proposta final após articular acordo entre três representações."')
    linha('Inválido: "Foi muito bom." / "Participou bem."')
    linha("Terminologia: usar Presidência (não usar Mesa); usar sessão, comitê, debates, discussões (não usar Plenária); usar representação e delegação.")

    return ws, triagem_ranges


# ---------------------------------------------------------------- Registro por Sessão

# colunas: A nº | B Sessão | C Delegado | D Representação/Função | E Presença |
# F Status | G-J eixo D | K-N eixo F | O-R eixo C | S-V eixo P | W X síntese
COL_EIXO = {"D": ("G", "H", "I", "J"), "F": ("K", "L", "M", "N"),
            "C": ("O", "P", "Q", "R"), "P": ("S", "T", "U", "V")}


def build_registro(wb, triagem_ranges):
    ws = wb.create_sheet(REG_NAME)
    ws.sheet_properties.tabColor = ROXO[2:]
    last_row = REG_DATA_START + N_DELEGADOS * N_SESSOES - 1

    title_block(ws, "X", "Controle Avaliação — Registro por Sessão", None)

    grupos = [("A2:F2", "Identificação")]
    for letra, nome, _, _ in EIXOS:
        cols = COL_EIXO[letra]
        grupos.append((f"{cols[0]}2:{cols[3]}2", nome))
    grupos.append(("W2:X2", "Síntese"))
    for rng, nome in grupos:
        ws.merge_cells(rng)
        first = rng.split(":")[0]
        c = ws[first]
        c.value = nome
        c.font = F_GRUPO
        c.alignment = AL_C
        col_ini = openpyxl.utils.column_index_from_string(first[:-1])
        col_fim = openpyxl.utils.column_index_from_string(rng.split(":")[1][:-1])
        for col in range(col_ini, col_fim + 1):
            ws.cell(row=2, column=col).fill = FILL_HDR
    ws.row_dimensions[2].height = 24

    headers = ["", "Sessão", "Delegado", "Representação/Função",
               "Presença na sessão", "Status do registro"]
    for letra, _, _, _ in EIXOS:
        headers += [f"Evidência {letra}", f"Gatilho {letra}",
                    f"Triagem {letra}", f"Nota Preliminar {letra}"]
    headers += ["Síntese acumulada da sessão", "Observação geral da sessão"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = AL_C
        c.border = BORDA_FINA
    ws.row_dimensions[3].height = 30

    bloco_inicio = {openpyxl.utils.column_index_from_string(COL_EIXO[l][0])
                    for l in COL_EIXO} | {openpyxl.utils.column_index_from_string("W")}

    for i in range(1, N_DELEGADOS + 1):
        rp = PREENCH_HDR_ROW + i  # linha do delegado no Preenchimento
        for s in range(1, N_SESSOES + 1):
            r = REG_DATA_START + (i - 1) * N_SESSOES + (s - 1)
            ws.cell(row=r, column=1, value=f'=IF($C{r}="","",{i})')
            ws.cell(row=r, column=2, value=f'=IF($C{r}="","","{s}ª Sessão")')
            ws.cell(row=r, column=3,
                    value=f'=IF(Preenchimento!$B${rp}="","",Preenchimento!$B${rp})')
            ws.cell(row=r, column=4,
                    value=f'=IF($C{r}="","",IF(Preenchimento!$C${rp}="","-",Preenchimento!$C${rp}))')
            for letra, _, _, base in EIXOS:
                cols = COL_EIXO[letra]
                nota_col = openpyxl.utils.column_index_from_string(cols[3])
                ws.cell(row=r, column=nota_col,
                        value=formula_nota_preliminar(letra, base, cols[1], r))

            fill = stripe_fill(i % 2 == 0)  # listra por bloco de delegado
            for j in range(1, 25):
                c = ws.cell(row=r, column=j)
                c.font = F_TXT_B if j in (10, 14, 18, 22) else F_TXT
                c.border = BORDA_BLOCO if j in bloco_inicio else BORDA_FINA
                if j in (7, 9, 11, 13, 15, 17, 19, 21, 23, 24):   # textos longos
                    c.alignment = AL_L
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                if fill:
                    c.fill = fill

    # bordas pretas de bloco também nos cabeçalhos
    for row in (2, 3):
        for col in sorted(bloco_inicio):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(left=LADO_PRETO, right=LADO_FINO,
                                 top=LADO_FINO, bottom=LADO_FINO)

    # validações
    dv_pres = DataValidation(type="list", formula1='"Presente,Falta"', allow_blank=True,
                             showErrorMessage=True, error="Use apenas Presente ou Falta.")
    ws.add_data_validation(dv_pres)
    dv_pres.add(f"E{REG_DATA_START}:E{last_row}")

    dv_status = DataValidation(type="list",
                               formula1=f'"{",".join(STATUS_REGISTRO)}"',
                               allow_blank=True, showErrorMessage=True,
                               error="Use um dos status previstos.")
    ws.add_data_validation(dv_status)
    dv_status.add(f"F{REG_DATA_START}:F{last_row}")

    for letra in COL_EIXO:
        col_triagem = COL_EIXO[letra][2]
        dv = DataValidation(type="list", formula1=triagem_ranges[letra],
                            allow_blank=True, showErrorMessage=True,
                            error="Use uma opção de Triagem da Rubrica.")
        ws.add_data_validation(dv)
        dv.add(f"{col_triagem}{REG_DATA_START}:{col_triagem}{last_row}")

    ws.conditional_formatting.add(
        f"A{REG_DATA_START}:X{last_row}",
        FormulaRule(formula=[f'$E{REG_DATA_START}="Falta"'], fill=FILL_FALTA))

    ws.freeze_panes = f"E{REG_DATA_START}"
    ws.auto_filter.ref = f"A3:X{last_row}"
    widths = {"A": 4, "B": 11, "C": 24, "D": 22, "E": 13, "F": 14, "W": 36, "X": 30}
    for letra in COL_EIXO:
        ev, ga, tr, no = COL_EIXO[letra]
        widths.update({ev: 38, ga: 12, tr: 26, no: 11})
    set_widths(ws, widths)
    return ws


# ---------------------------------------------------------------- Nota

# colunas: A nº | B Delegado | C Função | D Presença geral | E-H médias por eixo |
# I média geral | J evidências | K sínteses | L separador | M-AF colagem (20)
def build_nota(wb):
    ws = wb.create_sheet("Nota")
    ws.sheet_properties.tabColor = ROXO[2:]
    last_row = NOTA_DATA_START + N_DELEGADOS - 1

    ws.merge_cells("A1:K1")
    ws["A1"] = "Controle Avaliação — Nota"
    ws["A1"].font = F_TITULO
    ws["A1"].alignment = AL_L_NW
    ws.merge_cells("M1:AF1")
    ws["M1"] = ("Área de colagem — copiar M4:AF103 e colar somente valores em "
                "'Avaliações'!E4 da Planilha de Notas oficial")
    ws["M1"].font = F_SUBTITULO
    ws["M1"].alignment = AL_L_NW
    ws.row_dimensions[1].height = 30
    for col in range(1, 33):
        ws.cell(row=1, column=col).fill = FILL_TITULO

    ws.merge_cells("A2:K2")
    ws["A2"] = "Conferência e síntese (não colar na Planilha de Notas)"
    ws["A2"].font = F_SUBTITULO
    ws["A2"].alignment = AL_L_NW
    grupos = [("M2:Q2", EIXOS[0][2]), ("R2:V2", EIXOS[1][2]),
              ("W2:AA2", EIXOS[2][2]), ("AB2:AF2", EIXOS[3][2])]
    for rng, nome in grupos:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = nome
        c.font = F_HDR12
        c.alignment = AL_C
    for col in range(1, 33):
        ws.cell(row=2, column=col).fill = FILL_HDR
    ws.row_dimensions[2].height = 24

    headers = ["", "Delegado", "Representação/Função", "Presença geral",
               "Média D", "Média F", "Média C", "Média P", "Média geral",
               "Evidências das 5 sessões", "Sínteses e observações das 5 sessões", ""]
    headers += ["1ª Sessão", "2ª Sessão", "3ª Sessão", "4ª Sessão", "5ª Sessão"] * 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = AL_C
        c.border = BORDA_FINA
    ws.row_dimensions[3].height = 28

    # colunas de colagem: M..AF na ordem D1..D5 F1..F5 C1..C5 P1..P5
    paste_cols = [get_column_letter(13 + k) for k in range(20)]  # M=13
    nota_cols_reg = {"D": "J", "F": "N", "C": "R", "P": "V"}

    for i in range(1, N_DELEGADOS + 1):
        r = NOTA_DATA_START + i - 1
        rp = PREENCH_HDR_ROW + i
        reg0 = REG_DATA_START + (i - 1) * N_SESSOES  # linha da 1ª sessão no Registro

        ws.cell(row=r, column=1, value=f'=IF($B{r}="","",ROW()-{NOTA_DATA_START - 1})')
        ws.cell(row=r, column=2,
                value=f'=IF(Preenchimento!$B${rp}="","",Preenchimento!$B${rp})')
        ws.cell(row=r, column=3,
                value=f'=IF($B{r}="","",IF(Preenchimento!$C${rp}="","-",Preenchimento!$C${rp}))')
        ws.cell(row=r, column=4,
                value=f'=IF($B{r}="","",IF(Preenchimento!$D${rp}="","-",Preenchimento!$D${rp}))')
        ws.cell(row=r, column=5, value=f'=IF($B{r}="","",IFERROR(AVERAGE(M{r}:Q{r}),"-"))')
        ws.cell(row=r, column=6, value=f'=IF($B{r}="","",IFERROR(AVERAGE(R{r}:V{r}),"-"))')
        ws.cell(row=r, column=7, value=f'=IF($B{r}="","",IFERROR(AVERAGE(W{r}:AA{r}),"-"))')
        ws.cell(row=r, column=8, value=f'=IF($B{r}="","",IFERROR(AVERAGE(AB{r}:AF{r}),"-"))')
        ws.cell(row=r, column=9, value=f'=IF($B{r}="","",IFERROR(AVERAGE(M{r}:AF{r}),"-"))')

        # J: evidências das 5 sessões (rotuladas por sessão e eixo)
        termos = []
        for s in range(N_SESSOES):
            rr = reg0 + s
            for letra in ("D", "F", "C", "P"):
                ev_col = COL_EIXO[letra][0]
                ref = f"'{REG_NAME}'!${ev_col}{rr}"
                termos.append(f'IF({ref}="","","S{s + 1} {letra}: "&{ref}&CHAR(10))')
        ws.cell(row=r, column=10, value=f'=IF($B{r}="","",{"&".join(termos)})')

        # K: sínteses e observações das 5 sessões
        termos = []
        for s in range(N_SESSOES):
            rr = reg0 + s
            termos.append(f'IF(\'{REG_NAME}\'!$W{rr}="","","S{s + 1}: "&\'{REG_NAME}\'!$W{rr}&CHAR(10))')
            termos.append(f'IF(\'{REG_NAME}\'!$X{rr}="","","S{s + 1} obs: "&\'{REG_NAME}\'!$X{rr}&CHAR(10))')
        ws.cell(row=r, column=11, value=f'=IF($B{r}="","",{"&".join(termos)})')

        # M..AF: área de colagem (só números; falta/vazio vira célula vazia)
        for e, letra in enumerate(("D", "F", "C", "P")):
            reg_col = nota_cols_reg[letra]
            for s in range(N_SESSOES):
                col_letter = paste_cols[e * N_SESSOES + s]
                ref = f"'{REG_NAME}'!${reg_col}{reg0 + s}"
                ws.cell(row=r, column=openpyxl.utils.column_index_from_string(col_letter),
                        value=f'=IF(ISNUMBER({ref}),{ref},"")')

        fill = stripe_fill(i % 2 == 1)
        for j in range(1, 33):
            if j == 12:
                continue
            c = ws.cell(row=r, column=j)
            c.font = F_TXT
            c.border = BORDA_FINA
            if j in (10, 11):
                c.alignment = AL_L
            elif j in (2, 3):
                c.alignment = AL_L_NW
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if j in (5, 6, 7, 8, 9):
                c.number_format = "0.00"
            if fill:
                c.fill = fill

    # separador escuro entre conferência e colagem
    for row in range(1, last_row + 1):
        ws.cell(row=row, column=12).fill = FILL_TITULO
    ws.column_dimensions["L"].width = 2

    ws.conditional_formatting.add(
        f"A{NOTA_DATA_START}:K{last_row}",
        FormulaRule(formula=[f'$D{NOTA_DATA_START}="Falta"'], fill=FILL_FALTA))

    ws.freeze_panes = f"C{NOTA_DATA_START}"
    widths = {"A": 4, "B": 24, "C": 22, "D": 13, "E": 9, "F": 9, "G": 9, "H": 9,
              "I": 10, "J": 50, "K": 44}
    for col in paste_cols:
        widths[col] = 9.5
    set_widths(ws, widths)
    return ws


# ---------------------------------------------------------------- main

def main():
    # autoteste do parser de tokens antes de gerar
    casos = [("C+2", "C", 2), ("C+2; C-1", "C", 1), ("c+2;c-1", "C", 1),
             ("F+2; F-1", "F", 1), ("D+1; D-2", "D", -1), ("P-2", "P", -2),
             ("", "C", 0), ("obs", "C", 0), ("C+2, C-1", "C", 0),
             ("D+1;D+1;D+1;D+1;D+1;D+1;D+1;D+1", "D", 8), ("C+2; F-1", "C", 2)]
    for texto, letra, esperado in casos:
        obtido = simulate_tokens(texto, letra)
        assert obtido == esperado, f"parser: {texto!r} ({letra}) -> {obtido}, esperado {esperado}"
    print(f"Parser de tokens validado em {len(casos)} casos.")

    wb = openpyxl.Workbook()
    build_preenchimento(wb)
    _, triagem_ranges = build_rubrica(wb)
    build_registro(wb, triagem_ranges)
    build_nota(wb)
    out = "output/Controle Avaliação.xlsx"
    wb.save(out)
    print(f"Gerado: {out}")
    print("Abas:", wb.sheetnames)


if __name__ == "__main__":
    main()
